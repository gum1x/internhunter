from __future__ import annotations

import asyncio
import ipaddress
import math
import shutil
import socket
import zipfile
from collections.abc import Iterator
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any
from urllib.parse import urlsplit

import httpx
from sqlalchemy import select

from internhunter.config.settings import Settings, get_settings
from internhunter.core.db import (
    Company,
    DisclosureLead,
    get_session,
    init_db,
    upsert_disclosure_leads,
)
from internhunter.core.fetch import build_fetch_context
from internhunter.core.normalize import (
    canonical_company_slug,
    make_url_hash,
    parse_datetime,
)

_GENERIC_DOMAINS = {
    "gmail.com", "yahoo.com", "ymail.com", "outlook.com", "hotmail.com", "aol.com",
    "proton.me", "protonmail.com", "icloud.com", "me.com", "mac.com", "live.com",
    "msn.com", "gmx.com", "mail.com", "hey.com", "comcast.net", "verizon.net",
    "sbcglobal.net", "att.net",
}
_SENTINELS = {"", "na", "n/a", "n.a.", "none", "null", "nan", "#n/a", "(blank)", "-"}
_SBIR_DEFAULT = "https://api.www.sbir.gov/public/api/awards"
_MAX_ROWS = 2_000_000  # bound a decompression bomb / runaway sheet


@dataclass
class DisclosureSummary:
    rows: int = 0
    leads: int = 0
    companies: int = 0
    errors: list[str] = field(default_factory=list)


def _pick(row: dict[str, Any], *names: str) -> str | None:
    for name in names:
        value = row.get(name)
        if value is None:
            continue
        if isinstance(value, float) and math.isnan(value):
            continue
        text = str(value).strip()
        if text and text.lower() not in _SENTINELS:
            return text
    return None


def _valid_email(value: str | None) -> str | None:
    if not value:
        return None
    email = value.strip().lower()
    if "@" not in email or "." not in email.split("@", 1)[1]:
        return None
    return email


def _domain_from_email(email: str | None) -> str | None:
    if not email or "@" not in email:
        return None
    domain = email.split("@", 1)[1].strip().lower().rstrip(".")
    if not domain or domain in _GENERIC_DOMAINS:
        return None
    if "." not in domain or len(domain.rsplit(".", 1)[1]) < 2:
        return None
    return domain


def _full_name(first: str | None, last: str | None) -> str | None:
    parts = [p for p in (first, last) if p]
    return " ".join(parts) if parts else None


def _is_tech_soc(soc: str | None, prefixes: list[str]) -> bool:
    if not soc or not prefixes:
        return False
    code = soc.strip().replace(".", "-")
    return any(code.startswith(p) for p in prefixes)


def tech_company_of(row: dict[str, Any], soc_prefixes: list[str]) -> tuple[str, str] | None:
    """(employer_name, soc) if this filing is a tech-SOC hire, else None — used to count the
    per-company hiring SIGNAL even for rows that yield no contact (e.g. PERM/attorney-only)."""
    employer = _pick(row, "EMPLOYER_NAME", "EMPLOYER_BUSINESS_DBA")
    soc = _pick(row, "SOC_CODE", "PW_SOC_CODE", "LCA_CASE_SOC_CODE")
    if not employer or not _is_tech_soc(soc, soc_prefixes):
        return None
    return employer, soc or ""


def lead_from_lca_row(
    row: dict[str, Any], soc_prefixes: list[str], source: str = "oflc_lca"
) -> DisclosureLead | None:
    """Map one OFLC LCA/PERM disclosure row to a DisclosureLead, ONLY for the employer's own
    point-of-contact email. Third-party attorney/agent emails are deliberately NOT harvested
    (they are not hiring contacts and raise GDPR/relevance concerns); attorney-only filings
    still feed the company hiring signal via tech_company_of, just no personal contact."""
    tech = tech_company_of(row, soc_prefixes)
    if tech is None:
        return None
    employer, soc = tech

    poc_email = _valid_email(
        _pick(row, "EMPLOYER_POC_EMAIL", "EMPLOYER_POINT_OF_CONTACT_EMAIL")
    )
    if not poc_email:
        return None

    name = _full_name(
        _pick(row, "EMPLOYER_POC_FIRST_NAME", "EMPLOYER_POINT_OF_CONTACT_FIRST_NAME"),
        _pick(row, "EMPLOYER_POC_LAST_NAME", "EMPLOYER_POINT_OF_CONTACT_LAST_NAME"),
    )
    title = (
        _pick(row, "EMPLOYER_POC_JOB_TITLE", "EMPLOYER_POINT_OF_CONTACT_JOB_TITLE")
        or "HR Contact"
    )
    wage = _pick(
        row, "WAGE_RATE_OF_PAY_FROM_1", "WAGE_RATE_OF_PAY_FROM",
        "WAGE_OFFER_FROM_9089", "PW_AMOUNT_9089", "PREVAILING_WAGE",
    )
    return DisclosureLead(
        company_slug=canonical_company_slug(employer),
        company_name=employer,
        domain=_domain_from_email(poc_email),
        full_name=name,
        title=title,
        email=poc_email,
        phone=_pick(row, "EMPLOYER_POC_PHONE", "EMPLOYER_PHONE"),
        role_hint="hr",
        source=source,
        signal={
            "soc": soc,
            "soc_title": _pick(row, "SOC_TITLE", "PW_SOC_TITLE"),
            "job_title": _pick(row, "JOB_TITLE", "JOB_INFO_JOB_TITLE", "JOB_TITLE_1"),
            "wage": wage,
            "tech": True,
        },
        filed_at=parse_datetime(
            _pick(row, "RECEIVED_DATE", "CASE_RECEIVED_DATE", "DECISION_DATE")
        ),
    )


def leads_from_sbir_award(award: dict[str, Any]) -> list[DisclosureLead]:
    firm = _pick(award, "firm", "company")
    if not firm:
        return []
    slug = canonical_company_slug(firm)
    company_url = _pick(award, "company_url", "uei_url")
    domain: str | None = None
    if company_url:
        host = urlsplit(company_url if "//" in company_url else f"https://{company_url}").netloc
        domain = host.lower().removeprefix("www.").split(":")[0] or None
    signal = {
        "award": _pick(award, "award_amount"),
        "keywords": _pick(award, "research_area_keywords"),
        "employees": _pick(award, "number_employees"),
        "tech": True,
    }
    out: list[DisclosureLead] = []
    contacts = (
        ("poc_name", "poc_title", "poc_email", "poc_phone", "hr"),
        ("pi_name", "pi_title", "pi_email", "pi_phone", "hiring_manager"),
    )
    for name_key, title_key, email_key, phone_key, role in contacts:
        email = _valid_email(_pick(award, email_key))
        name = _pick(award, name_key)
        if not email and not name:
            continue
        out.append(
            DisclosureLead(
                company_slug=slug,
                company_name=firm,
                domain=domain or _domain_from_email(email),
                full_name=name,
                title=_pick(award, title_key)
                or ("Principal Investigator" if role == "hiring_manager" else "Company Contact"),
                email=email,
                phone=_pick(award, phone_key),
                role_hint=role,
                source="sbir",
                signal=signal,
            )
        )
    return out


def _dedup_headers(raw: tuple[Any, ...]) -> list[str]:
    counts: dict[str, int] = {}
    out: list[str] = []
    for i, cell in enumerate(raw):
        base = (str(cell).strip() if cell is not None else "") or f"_col{i}"
        seen = counts.get(base, 0)
        counts[base] = seen + 1
        out.append(base if seen == 0 else f"{base}__{seen}")
    return out


def _iter_xlsx_rows(path: Path) -> Iterator[dict[str, Any]]:
    from openpyxl import load_workbook

    workbook = load_workbook(filename=path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        if sheet is None:
            return
        rows = sheet.iter_rows(values_only=True)
        try:
            header = _dedup_headers(next(rows))
        except StopIteration:
            return
        width = len(header)
        for values in rows:
            yield {header[i]: (values[i] if i < len(values) else None) for i in range(width)}
    finally:
        workbook.close()


def _validate_remote_url(url: str) -> str | None:
    parts = urlsplit(url)
    if parts.scheme not in ("http", "https"):
        return f"unsupported URL scheme: {parts.scheme!r}"
    host = parts.hostname
    if not host:
        return "URL has no host"
    try:
        infos = socket.getaddrinfo(host, None)
    except OSError as exc:
        return f"could not resolve {host}: {exc}"
    for info in infos:
        ip = ipaddress.ip_address(info[4][0])
        if not ip.is_global:
            return f"refusing to fetch non-public host {host} ({ip})"
    return None


def _ensure_xlsx(path: Path, cache_dir: Path) -> Path:
    """A real DOL download is often a .zip wrapping the .xlsx; an .xlsx is itself a zip but
    contains no nested .xlsx. Extract the nested workbook if present, else use the file."""
    if not zipfile.is_zipfile(path):
        return path
    with zipfile.ZipFile(path) as zf:
        members = [n for n in zf.namelist() if n.lower().endswith(".xlsx")]
        if not members:
            return path
        member = max(members, key=lambda n: zf.getinfo(n).file_size)
        out = cache_dir / f"{path.name}.extracted.xlsx"
        with zf.open(member) as src, out.open("wb") as dst:
            shutil.copyfileobj(src, dst)
    return out


async def _download(settings: Settings, url: str) -> Path:
    error = await asyncio.to_thread(_validate_remote_url, url)
    if error:
        raise ValueError(error)
    user_agent = settings.disclosure_user_agent or settings.default_user_agent
    settings.cache_dir.mkdir(parents=True, exist_ok=True)
    raw = settings.cache_dir / f"disclosure_{make_url_hash(url)}"
    timeout = httpx.Timeout(connect=30.0, read=300.0, write=300.0, pool=30.0)
    async with httpx.AsyncClient(
        headers={"User-Agent": user_agent},
        follow_redirects=True,
        timeout=timeout,
        proxy=settings.http_proxy or None,
    ) as client:
        async with client.stream("GET", url) as response:
            response.raise_for_status()
            with raw.open("wb") as handle:
                async for chunk in response.aiter_bytes(1 << 20):
                    handle.write(chunk)
    return _ensure_xlsx(raw, settings.cache_dir)


def _apply_company_signals(slug_signal: dict[str, dict[str, Any]]) -> int:
    """Merge each canonical-slug hiring signal onto an EXISTING company (matched by the
    company's own canonical slug/name). Enrichment-only: we never create net-new 'pending'
    rows (select_companies only targets companies with internship jobs, so they'd be inert);
    the leads stay in disclosure_leads and are consumed once the company is polled."""
    if not slug_signal:
        return 0
    session = get_session()
    touched = 0
    try:
        by_canon: dict[str, list[Company]] = {}
        for company in session.scalars(select(Company)):
            key = canonical_company_slug(company.name or company.company_slug)
            if key:
                by_canon.setdefault(key, []).append(company)
        for canon, signal in slug_signal.items():
            for company in by_canon.get(canon, []):
                existing = (company.notes or {}).get("disclosure", {})
                merged = {**existing, **{k: v for k, v in signal.items() if k != "domain"}}
                company.notes = {**(company.notes or {}), "disclosure": merged}
                if signal.get("domain") and not company.domain:
                    company.domain = signal["domain"]
                touched += 1
        session.commit()
    finally:
        session.close()
    return touched


def _bump_signal(
    slug_signal: dict[str, dict[str, Any]], slug: str, key: str, name: str | None, **extra: Any
) -> None:
    entry = slug_signal.setdefault(slug, {})
    entry[key] = int(entry.get(key, 0)) + 1
    if name:
        entry["name"] = name
    entry.update(extra)


async def ingest_oflc(
    settings: Settings | None = None,
    *,
    source: str | Path | None = None,
    program: str = "oflc_lca",
) -> DisclosureSummary:
    resolved = settings or get_settings()
    init_db(resolved.db_path)
    summary = DisclosureSummary()
    target = source or resolved.oflc_lca_url
    if not target:
        summary.errors.append(
            "no OFLC source: pass --url (a data.gov LCA .xlsx/.zip URL or local path) "
            "or set INTERNHUNTER_OFLC_LCA_URL"
        )
        return summary

    if isinstance(target, str) and target.startswith("http"):
        try:
            path = await _download(resolved, target)
        except Exception as exc:
            summary.errors.append(f"download failed: {exc}")
            return summary
    else:
        path = Path(target)
        if not path.exists():
            summary.errors.append(f"file not found: {path}")
            return summary

    prefixes = [p.strip() for p in resolved.oflc_soc_prefixes.split(",") if p.strip()]
    leads: list[DisclosureLead] = []
    slug_signal: dict[str, dict[str, Any]] = {}
    try:
        for row in _iter_xlsx_rows(path):
            summary.rows += 1
            if summary.rows > _MAX_ROWS:
                summary.errors.append(f"row cap {_MAX_ROWS} reached; truncating")
                break
            tech = tech_company_of(row, prefixes)
            if tech is None:
                continue
            employer, soc = tech
            lead = lead_from_lca_row(row, prefixes, source=program)
            slug = lead.company_slug if lead else canonical_company_slug(employer)
            _bump_signal(slug_signal, slug, "tech_filings", employer, last_soc=soc)
            if lead is not None:
                slug_signal[slug]["domain"] = lead.domain or slug_signal[slug].get("domain")
                leads.append(lead)
    except Exception as exc:
        summary.errors.append(f"parse failed: {exc}")
        return summary

    session = get_session()
    try:
        summary.leads = upsert_disclosure_leads(session, leads)
    finally:
        session.close()
    summary.companies = _apply_company_signals(slug_signal)
    return summary


async def ingest_sbir(
    settings: Settings | None = None, *, year: int | None = None, rows: int = 500
) -> DisclosureSummary:
    resolved = settings or get_settings()
    init_db(resolved.db_path)
    summary = DisclosureSummary()
    params: dict[str, Any] = {"rows": rows}
    if year is not None:
        params["year"] = year
    user_agent = resolved.disclosure_user_agent or resolved.default_user_agent
    try:
        async with build_fetch_context(resolved) as ctx:
            data = await ctx.get_json(
                resolved.sbir_api_url or _SBIR_DEFAULT,
                params=params,
                headers={"User-Agent": user_agent},
                respect_robots=False,
            )
    except Exception as exc:
        summary.errors.append(f"sbir fetch failed: {exc}")
        return summary

    awards = data if isinstance(data, list) else (data.get("results") or data.get("data") or [])
    leads: list[DisclosureLead] = []
    slug_signal: dict[str, dict[str, Any]] = {}
    for award in awards:
        if not isinstance(award, dict):
            continue
        summary.rows += 1
        award_leads = leads_from_sbir_award(award)
        if award_leads:
            slug = award_leads[0].company_slug
            _bump_signal(
                slug_signal, slug, "sbir_awards", award_leads[0].company_name,
                domain=next((lead.domain for lead in award_leads if lead.domain), None),
            )
            leads.extend(award_leads)

    session = get_session()
    try:
        summary.leads = upsert_disclosure_leads(session, leads)
    finally:
        session.close()
    summary.companies = _apply_company_signals(slug_signal)
    return summary


def run_ingest_disclosure(
    source: str, settings: Settings | None = None, url: str | None = None
) -> DisclosureSummary:
    if source == "sbir":
        return asyncio.run(ingest_sbir(settings))
    if source == "perm":
        return asyncio.run(ingest_oflc(settings, source=url, program="oflc_perm"))
    return asyncio.run(ingest_oflc(settings, source=url, program="oflc_lca"))
